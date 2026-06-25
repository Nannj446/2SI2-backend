"""
Servicios de exportación para reportes manuales.

Maneja exportación a CSV, XLSX y PDF con:
- Encabezados en español
- Formato profesional
- Gráficos incluidos

Autor: Sistema FinCore
Fecha: 2026-05-11
"""

import csv
import io
from datetime import datetime
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart

from .column_labels import get_column_label, translate_row_keys, get_ordered_columns


class ManualExportService:
    """
    Servicio para exportar reportes manuales en diferentes formatos.
    """
    
    def __init__(self, data, report_type, filters):
        """
        Inicializa el servicio de exportación.
        
        Args:
            data: Datos del reporte
            report_type: Tipo de reporte
            filters: Filtros aplicados
        """
        self.data = data
        self.report_type = report_type
        self.filters = filters
    
    # ============================================================
    # EXPORTACIÓN CSV
    # ============================================================
    
    def export_csv(self):
        """
        Exporta el reporte a formato CSV con encabezados en español.
        
        Returns:
            HttpResponse: Respuesta con archivo CSV
        """
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"reporte-{self.report_type}-{datetime.now().strftime('%Y%m%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Agregar BOM para Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Obtener filas
        rows = self.data.get('rows', [])
        if not rows:
            return response
        
        # Obtener columnas ordenadas
        ordered_columns = get_ordered_columns(self.report_type)
        if not ordered_columns:
            ordered_columns = list(rows[0].keys())
        
        # Escribir headers traducidos
        headers = [get_column_label(self.report_type, col) for col in ordered_columns]
        writer.writerow(headers)
        
        # Escribir datos
        for row in rows:
            writer.writerow([self._format_value(row.get(key, '')) for key in ordered_columns])
        
        return response
    
    # ============================================================
    # EXPORTACIÓN XLSX
    # ============================================================
    
    def export_xlsx(self):
        """
        Exporta el reporte a formato Excel (XLSX) con encabezados en español.
        
        Returns:
            HttpResponse: Respuesta con archivo XLSX
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.chart import PieChart, BarChart, Reference
        except ImportError:
            # Fallback a CSV si openpyxl no está disponible
            return self.export_csv()
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos"
        
        # Obtener filas
        rows = self.data.get('rows', [])
        
        if not rows:
            # Crear respuesta vacía
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"reporte-{self.report_type}-{datetime.now().strftime('%Y%m%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
        
        # Obtener columnas ordenadas
        ordered_columns = get_ordered_columns(self.report_type)
        if not ordered_columns:
            ordered_columns = list(rows[0].keys())
        
        # Headers traducidos
        headers = [get_column_label(self.report_type, col) for col in ordered_columns]
        ws.append(headers)
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar estilo a headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        for row in rows:
            row_values = [self._format_value(row.get(key, '')) for key in ordered_columns]
            ws.append(row_values)
        
        # Aplicar bordes y alineación a datos
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center')
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Agregar hoja de resumen si hay datos
        summary = self.data.get('summary', {})
        if summary:
            ws_summary = wb.create_sheet("Resumen", 0)
            self._add_summary_sheet(ws_summary, summary)
        
        # Crear respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"reporte-{self.report_type}-{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    def _add_summary_sheet(self, ws, summary):
        """Agrega hoja de resumen al Excel."""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Título
        ws['A1'] = self.get_report_title()
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:B1')
        
        # Fecha
        ws['A2'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:B2')
        
        # Resumen
        row = 4
        ws[f'A{row}'] = "Resumen"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        # Datos del resumen
        for key, value in summary.items():
            if isinstance(value, (int, float, str)):
                label = self._format_summary_key(key)
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    # ============================================================
    # EXPORTACIÓN PDF
    # ============================================================
    
    def export_pdf(self, include_chart=False):
        """
        Exporta el reporte a formato PDF con gráficos.
        
        Args:
            include_chart: Si incluir gráficos
        
        Returns:
            HttpResponse: Respuesta con archivo PDF
        """
        response = HttpResponse(content_type='application/pdf')
        filename = f"reporte-{self.report_type}-{datetime.now().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Usar landscape para más espacio
        doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                               rightMargin=30, leftMargin=30,
                               topMargin=30, bottomMargin=30)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#366092'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            fontName='Helvetica-Bold'
        )
        
        # Título
        title = Paragraph(self.get_report_title(), title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Información del reporte
        info_text = f"<b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        info = Paragraph(info_text, styles['Normal'])
        elements.append(info)
        elements.append(Spacer(1, 0.3*inch))
        
        # Resumen
        summary = self.data.get('summary', {})
        if summary:
            summary_title = Paragraph("Resumen Ejecutivo", subtitle_style)
            elements.append(summary_title)
            elements.append(Spacer(1, 0.1*inch))
            
            summary_text = self.format_summary_for_pdf(summary)
            summary_para = Paragraph(summary_text, styles['Normal'])
            elements.append(summary_para)
            elements.append(Spacer(1, 0.3*inch))
        
        # Gráficos
        if include_chart:
            chart_data = self.data.get('chart_data', {})
            if chart_data:
                chart_title = Paragraph("Visualización de Datos", subtitle_style)
                elements.append(chart_title)
                elements.append(Spacer(1, 0.2*inch))
                
                # Agregar gráfico principal
                chart = self._create_chart_for_pdf(chart_data)
                if chart:
                    elements.append(chart)
                    elements.append(Spacer(1, 0.4*inch))
                else:
                    # Si no hay gráfico, agregar nota
                    no_chart_note = Paragraph(
                        "<i>No hay datos suficientes para generar gráficos.</i>",
                        styles['Normal']
                    )
                    elements.append(no_chart_note)
                    elements.append(Spacer(1, 0.3*inch))
        
        # Tabla de datos
        rows = self.data.get('rows', [])
        if rows:
            data_title = Paragraph("Datos Detallados", subtitle_style)
            elements.append(data_title)
            elements.append(Spacer(1, 0.1*inch))
            
            # Usar todas las filas para PDF (sin límite)
            limited_rows = rows
            
            # Obtener columnas ordenadas
            ordered_columns = get_ordered_columns(self.report_type)
            if not ordered_columns:
                ordered_columns = list(limited_rows[0].keys())
            
            # Limitar columnas para que quepan en la página
            max_columns = 8
            if len(ordered_columns) > max_columns:
                ordered_columns = ordered_columns[:max_columns]
            
            # Preparar datos de tabla con headers traducidos
            headers = [get_column_label(self.report_type, col) for col in ordered_columns]
            table_data = [headers]
            
            for row in limited_rows:
                row_data = []
                for key in ordered_columns:
                    value = self._format_value(row.get(key, ''))
                    # Limitar longitud del texto para que quepa en la celda
                    str_value = str(value)
                    if len(str_value) > 30:
                        str_value = str_value[:27] + '...'
                    row_data.append(str_value)
                table_data.append(row_data)
            
            # Calcular ancho de columnas dinámicamente
            available_width = landscape(A4)[0] - 60  # Restar márgenes
            col_width = available_width / len(headers)
            
            # Ajustar anchos según el tipo de columna
            col_widths = []
            for header in headers:
                # Columnas más estrechas para IDs y códigos
                if any(x in header.lower() for x in ['id', 'código']):
                    col_widths.append(col_width * 0.6)
                # Columnas más anchas para descripciones
                elif any(x in header.lower() for x in ['descripción', 'nombre', 'cliente']):
                    col_widths.append(col_width * 1.2)
                else:
                    col_widths.append(col_width)
            
            # Normalizar anchos para que sumen el ancho disponible
            total_width = sum(col_widths)
            col_widths = [w * available_width / total_width for w in col_widths]
            
            # Crear tabla
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                # Datos
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 6),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                # Ajuste de texto
                ('WORDWRAP', (0, 0), (-1, -1), True),
            ]))
            
            elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        return response
    
    def _create_chart_for_pdf(self, chart_data):
        """
        Crea un gráfico para incluir en el PDF.
        
        Args:
            chart_data: Datos del gráfico
        
        Returns:
            Drawing: Gráfico de ReportLab o None
        """
        try:
            # Seleccionar el gráfico más relevante según el tipo de reporte
            if self.report_type == 'clients':
                # Prioridad: estado > tipo > otros
                if 'by_status' in chart_data and chart_data['by_status']:
                    return self._create_pie_chart(chart_data['by_status'], "Distribución de Clientes por Estado")
                elif 'by_type' in chart_data and chart_data['by_type']:
                    return self._create_pie_chart(chart_data['by_type'], "Distribución de Clientes por Tipo")
                    
            elif self.report_type == 'products':
                # Prioridad: tipo > estado
                if 'by_type' in chart_data and chart_data['by_type']:
                    return self._create_pie_chart(chart_data['by_type'], "Distribución de Productos por Tipo")
                elif 'by_status' in chart_data and chart_data['by_status']:
                    return self._create_pie_chart(chart_data['by_status'], "Distribución de Productos por Estado")
                    
            elif self.report_type == 'applications':
                # Prioridad: estado (barras) > otros
                if 'by_status' in chart_data and chart_data['by_status']:
                    return self._create_bar_chart(chart_data['by_status'], "Solicitudes por Estado")
                elif 'by_product' in chart_data and chart_data['by_product']:
                    return self._create_bar_chart(chart_data['by_product'], "Solicitudes por Producto")
                    
            elif self.report_type == 'audit':
                # Prioridad: severidad > acción
                if 'by_severity' in chart_data and chart_data['by_severity']:
                    return self._create_pie_chart(chart_data['by_severity'], "Distribución de Eventos por Severidad")
                elif 'by_action' in chart_data and chart_data['by_action']:
                    return self._create_bar_chart(chart_data['by_action'], "Eventos por Acción")
                    
            elif self.report_type == 'users':
                # Prioridad: rol > estado
                if 'by_role' in chart_data and chart_data['by_role']:
                    return self._create_pie_chart(chart_data['by_role'], "Distribución de Usuarios por Rol")
                elif 'by_status' in chart_data and chart_data['by_status']:
                    return self._create_pie_chart(chart_data['by_status'], "Distribución de Usuarios por Estado")
                    
            elif self.report_type == 'branches':
                # Prioridad: estado > ubicación
                if 'by_status' in chart_data and chart_data['by_status']:
                    return self._create_pie_chart(chart_data['by_status'], "Distribución de Sucursales por Estado")
                elif 'by_city' in chart_data and chart_data['by_city']:
                    return self._create_bar_chart(chart_data['by_city'], "Sucursales por Ciudad")
                    
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error creando gráfico para {self.report_type}: {str(e)}")
        
        return None
    
    def _create_pie_chart(self, data, title):
        """Crea un gráfico de pastel profesional con leyenda."""
        from reportlab.graphics.shapes import String, Rect
        
        drawing = Drawing(500, 250)
        
        pie = Pie()
        pie.x = 80
        pie.y = 80
        pie.width = 120
        pie.height = 120
        
        # Preparar datos
        labels = [item['name'] for item in data if item.get('value', 0) > 0]
        values = [item['value'] for item in data if item.get('value', 0) > 0]
        
        if not values:
            return None
        
        pie.data = values
        # No mostrar labels en el gráfico, usaremos leyenda
        pie.labels = None
        pie.slices.strokeWidth = 1
        pie.slices.strokeColor = colors.white
        
        # Colores profesionales
        colors_list = [
            colors.HexColor('#366092'),
            colors.HexColor('#52B788'),
            colors.HexColor('#F4A261'),
            colors.HexColor('#E76F51'),
            colors.HexColor('#2A9D8F'),
            colors.HexColor('#E9C46A'),
            colors.HexColor('#8E44AD'),
            colors.HexColor('#3498DB'),
        ]
        
        for i, color in enumerate(colors_list[:len(values)]):
            pie.slices[i].fillColor = color
        
        drawing.add(pie)
        
        # Título
        title_string = String(250, 230, title, textAnchor='middle', fontSize=14, fontName='Helvetica-Bold')
        drawing.add(title_string)
        
        # Leyenda a la derecha
        legend_x = 250
        legend_y = 200
        legend_spacing = 20
        
        for i, (label, value) in enumerate(zip(labels, values)):
            y_pos = legend_y - (i * legend_spacing)
            
            # Cuadro de color
            rect = Rect(legend_x, y_pos - 8, 12, 12)
            rect.fillColor = colors_list[i % len(colors_list)]
            rect.strokeColor = colors.grey
            rect.strokeWidth = 0.5
            drawing.add(rect)
            
            # Texto de la leyenda con valor
            percentage = (value / sum(values)) * 100
            legend_text = f"{label}: {value} ({percentage:.1f}%)"
            # Truncar si es muy largo
            if len(legend_text) > 35:
                legend_text = legend_text[:32] + "..."
            
            text = String(legend_x + 18, y_pos - 6, legend_text, 
                         fontSize=9, fontName='Helvetica')
            drawing.add(text)
        
        return drawing
    
    def _create_bar_chart(self, data, title):
        """Crea un gráfico de barras profesional."""
        from reportlab.graphics.shapes import String
        
        drawing = Drawing(500, 280)
        
        bc = VerticalBarChart()
        bc.x = 60
        bc.y = 60
        bc.height = 160
        bc.width = 380
        
        # Preparar datos
        labels = [item['name'][:20] for item in data if item.get('value', 0) > 0]
        values = [[item['value']] for item in data if item.get('value', 0) > 0]
        
        if not values:
            return None
        
        bc.data = values
        bc.categoryAxis.categoryNames = labels
        
        # Configuración del eje de categorías (X)
        bc.categoryAxis.labels.angle = 30
        bc.categoryAxis.labels.fontSize = 8
        bc.categoryAxis.labels.fontName = 'Helvetica'
        bc.categoryAxis.labels.dx = 0
        bc.categoryAxis.labels.dy = -5
        bc.categoryAxis.strokeWidth = 1
        bc.categoryAxis.strokeColor = colors.grey
        
        # Configuración del eje de valores (Y)
        bc.valueAxis.valueMin = 0
        bc.valueAxis.valueStep = None  # Auto
        bc.valueAxis.labels.fontSize = 8
        bc.valueAxis.labels.fontName = 'Helvetica'
        bc.valueAxis.strokeWidth = 1
        bc.valueAxis.strokeColor = colors.grey
        bc.valueAxis.gridStrokeColor = colors.lightgrey
        bc.valueAxis.gridStrokeWidth = 0.5
        
        # Estilo de las barras
        bc.bars[0].fillColor = colors.HexColor('#366092')
        bc.bars[0].strokeColor = colors.HexColor('#2A4A6F')
        bc.bars[0].strokeWidth = 1
        
        # Ancho de las barras
        bc.barWidth = 15
        bc.groupSpacing = 10
        
        drawing.add(bc)
        
        # Título
        title_string = String(250, 250, title, textAnchor='middle', fontSize=14, fontName='Helvetica-Bold')
        drawing.add(title_string)
        
        # Etiqueta del eje Y
        y_label = String(20, 140, 'Cantidad', textAnchor='middle', fontSize=10, fontName='Helvetica-Bold')
        y_label.textTransform = 'rotate(270)'
        drawing.add(y_label)
        
        return drawing
    
    # ============================================================
    # UTILIDADES
    # ============================================================
    
    def get_report_title(self):
        """Obtiene el título del reporte en español."""
        titles = {
            'clients': 'Reporte de Clientes',
            'products': 'Reporte de Productos Crediticios',
            'applications': 'Reporte de Solicitudes de Crédito',
            'audit': 'Reporte de Auditoría',
            'users': 'Reporte de Usuarios',
            'branches': 'Reporte de Sucursales',
        }
        return titles.get(self.report_type, 'Reporte')
    
    def format_summary_for_pdf(self, summary):
        """Formatea el resumen para PDF con etiquetas en español."""
        lines = []
        
        # Mapeo de claves a etiquetas en español
        labels = {
            'total': 'Total de registros',
            'active': 'Activos',
            'inactive': 'Inactivos',
            'verified': 'Verificados',
            'approval_rate': 'Tasa de aprobación',
            'avg_income': 'Ingreso promedio',
            'total_requested': 'Monto total solicitado',
            'total_approved': 'Monto total aprobado',
            'avg_processing_days': 'Días promedio de procesamiento',
        }
        
        for key, value in summary.items():
            if isinstance(value, (int, float, str)) and key in labels:
                label = labels[key]
                formatted_value = self._format_value(value)
                lines.append(f"<b>{label}:</b> {formatted_value}")
        
        return '<br/>'.join(lines) if lines else 'Sin datos de resumen'
    
    def _format_summary_key(self, key):
        """Formatea una clave del resumen a español."""
        labels = {
            'total': 'Total',
            'active': 'Activos',
            'inactive': 'Inactivos',
            'verified': 'Verificados',
            'approval_rate': 'Tasa de Aprobación (%)',
            'avg_income': 'Ingreso Promedio (Bs)',
            'total_requested': 'Monto Total Solicitado (Bs)',
            'total_approved': 'Monto Total Aprobado (Bs)',
            'avg_processing_days': 'Días Promedio de Procesamiento',
        }
        return labels.get(key, key.replace('_', ' ').title())
    
    def _format_value(self, value):
        """Formatea un valor para exportación."""
        if value is None:
            return ''
        if isinstance(value, bool):
            return 'Sí' if value else 'No'
        if isinstance(value, float):
            # Formatear números con 2 decimales
            return f"{value:,.2f}"
        if isinstance(value, str):
            # Traducir valores comunes
            translations = {
                'True': 'Sí',
                'False': 'No',
                'ACTIVE': 'Activo',
                'INACTIVE': 'Inactivo',
                'PENDING': 'Pendiente',
                'VERIFIED': 'Verificado',
                'REJECTED': 'Rechazado',
                'APPROVED': 'Aprobado',
                'DRAFT': 'Borrador',
                'SUBMITTED': 'Enviado',
                'IN_REVIEW': 'En Revisión',
                'DISBURSED': 'Desembolsado',
                'CANCELLED': 'Cancelado',
                'LOW': 'Bajo',
                'MEDIUM': 'Medio',
                'HIGH': 'Alto',
                'VERY_HIGH': 'Muy Alto',
                'FIXED': 'Fija',
                'VARIABLE': 'Variable',
                'MIXED': 'Mixta',
            }
            return translations.get(value, value)
        return str(value)
