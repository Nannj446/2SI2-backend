"""
Servicio de exportación de reportes.

Convierte datos de reportes a formatos de archivo:
- CSV
- XLSX (Excel)
- PDF
"""
import csv
import io
from typing import List, Dict, Any, Optional
from datetime import datetime, date
from decimal import Decimal

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


class ExportService:
    """
    Servicio de exportación de reportes.
    
    Convierte datos de reportes a formatos descargables.
    Soporta CSV, XLSX y PDF con formato profesional.
    """
    
    SUPPORTED_FORMATS = ['csv', 'xlsx', 'pdf']
    
    # Mapeo de nombres técnicos a nombres en español
    COLUMN_LABELS = {
        # Información de solicitud
        'application_number': 'Número de Solicitud',
        'status': 'Estado',
        'risk_level': 'Nivel de Riesgo',
        'credit_score': 'Puntaje de Crédito',
        'purpose': 'Propósito del Crédito',
        'notes': 'Notas',
        'internal_notes': 'Notas Internas',
        'observation_reason': 'Motivo de Observación',
        'rejection_reason': 'Motivo de Rechazo',
        'is_active': 'Activo',
        
        # Cliente
        'client_name': 'Nombre del Cliente',
        'client_document': 'Documento del Cliente',
        'client_email': 'Correo Electrónico',
        'client_phone': 'Teléfono',
        'client_type': 'Tipo de Cliente',
        'full_name': 'Nombre Completo',
        'first_name': 'Nombre',
        'last_name': 'Apellido',
        'document_number': 'Número de Documento',
        'document_type': 'Tipo de Documento',
        'document_extension': 'Extensión CI',
        'birth_date': 'Fecha de Nacimiento',
        'gender': 'Género',
        'email': 'Correo Electrónico',
        'phone': 'Teléfono',
        'mobile_phone': 'Teléfono Móvil',
        'address': 'Dirección',
        'city': 'Ciudad',
        'department': 'Departamento',
        'country': 'País',
        'postal_code': 'Código Postal',
        'kyc_status': 'Estado KYC',
        'employment_status': 'Estado Laboral',
        'employer_name': 'Nombre del Empleador',
        'employer_nit': 'NIT del Empleador',
        'job_title': 'Cargo',
        'employment_start_date': 'Fecha de Inicio Laboral',
        'monthly_income': 'Ingreso Mensual',
        'additional_income': 'Ingresos Adicionales',
        'last_activity_at': 'Última Actividad',
        'last_login': 'Último Acceso',
        'active_time': 'Tiempo Activo',
        'device_type': 'Tipo de Dispositivo',
        'verified_at': 'Fecha de Verificación',
        'verified_by_name': 'Verificado Por',
        
        # Producto
        'product_name': 'Producto',
        'product_code': 'Código de Producto',
        'product_type': 'Tipo de Producto',
        
        # Montos
        'requested_amount': 'Monto Solicitado',
        'approved_amount': 'Monto Aprobado',
        'disbursed_amount': 'Monto Desembolsado',
        'total_requested_amount': 'Total Solicitado',
        'total_approved_amount': 'Total Aprobado',
        'avg_requested_amount': 'Promedio Solicitado',
        'avg_approved_amount': 'Promedio Aprobado',
        
        # Plazos y tasas
        'term_months': 'Plazo (Meses)',
        'approved_term_months': 'Plazo Aprobado (Meses)',
        'interest_rate': 'Tasa de Interés',
        'approved_interest_rate': 'Tasa de Interés Aprobada',
        'monthly_payment': 'Cuota Mensual',
        'debt_to_income_ratio': 'Ratio Deuda/Ingreso',
        'avg_term_months': 'Promedio Plazo (Meses)',
        
        # Sucursal
        'branch_name': 'Sucursal',
        'branch_city': 'Ciudad',
        'branch_count': 'Cantidad de Sucursales',
        
        # Usuario asignado
        'assigned_to_name': 'Asignado a',
        'reviewed_by_name': 'Revisado Por',
        'approved_by_name': 'Aprobado Por',
        'created_by_name': 'Creado Por',
        'updated_by_name': 'Actualizado Por',
        
        # Estados de verificación
        'identity_verification_status': 'Estado de Verificación de Identidad',
        'documents_status': 'Estado de Documentos',
        'employment_type': 'Tipo de Empleo',
        
        # Fechas
        'created_at': 'Fecha de Creación',
        'submitted_at': 'Fecha de Envío',
        'reviewed_at': 'Fecha de Revisión',
        'approved_at': 'Fecha de Aprobación',
        'rejected_at': 'Fecha de Rechazo',
        'disbursed_at': 'Fecha de Desembolso',
        'updated_at': 'Fecha de Actualización',
        
        # Contadores
        'total_applications': 'Total de Solicitudes',
        'approved_count': 'Aprobadas',
        'rejected_count': 'Rechazadas',
        'pending_count': 'Pendientes',
        'approval_rate': 'Tasa de Aprobación',
        'total_active_loans': 'Total de Créditos Activos',
        'avg_credit_score': 'Puntaje de Crédito Promedio',
        'latest_loan_date': 'Fecha del Último Crédito',
        
        # Tenant
        'tenant_name': 'Institución',
        'tenant_slug': 'Código de Institución',
        'institution_type': 'Tipo de Institución',
        'subscription_status': 'Estado de Suscripción',
        'user_count': 'Cantidad de Usuarios',
        'total_clients': 'Total de Clientes',
        'active_loans_count': 'Créditos Activos',
        'total_users': 'Total de Usuarios',
        'active_users': 'Usuarios Activos',
        'inactive_users': 'Usuarios Inactivos',
        'admin_count': 'Administradores',
        'manager_count': 'Gerentes',
        'analyst_count': 'Analistas',
        'officer_count': 'Oficiales',
        'client_count': 'Clientes',
        'last_user_created_at': 'Último Usuario Creado',
        
        # Plan y suscripción
        'plan_name': 'Plan',
        'payment_status': 'Estado de Pago',
        'start_date': 'Fecha de Inicio',
        'end_date': 'Fecha de Fin',
        'trial_end_date': 'Fin de Período de Prueba',
        'next_billing_date': 'Próxima Fecha de Facturación',
        'amount_due': 'Monto a Pagar',
        'total_paid': 'Total Pagado',
        'current_users': 'Usuarios Actuales',
        'current_branches': 'Sucursales Actuales',
        'days_active': 'Días Activo',
        
        # Documentos
        'document_type': 'Tipo de Documento',
        'document_status': 'Estado del Documento',
        'total_documents_required': 'Total de Documentos Requeridos',
        'pending_documents_count': 'Documentos Pendientes',
        'pending_document_types': 'Tipos de Documentos Pendientes',
        'completion_percentage': 'Porcentaje de Completitud',
        'application_status': 'Estado de Solicitud',
        'days_since_submission': 'Días desde Envío',
        
        # Verificación
        'verification_status': 'Estado de Verificación',
        'verification_method': 'Método de Verificación',
        'decision': 'Decisión',
        'provider': 'Proveedor',
        'started_at': 'Fecha de Inicio',
        'completed_at': 'Fecha de Finalización',
        'processing_time_minutes': 'Tiempo de Procesamiento (minutos)',
    }
    
    # Títulos de reportes en español
    REPORT_TITLES = {
        'loans_by_product': 'Reporte de Créditos por Producto',
        'loans_by_branch': 'Reporte de Créditos por Sucursal',
        'loans_by_status': 'Reporte de Créditos por Estado',
        'loans_by_date_range': 'Reporte de Créditos por Rango de Fechas',
        'active_loans': 'Reporte de Créditos Activos',
        'customers_registered': 'Reporte de Clientes Registrados',
        'customers_by_status': 'Reporte de Clientes por Estado',
        'customers_with_active_loans': 'Reporte de Clientes con Créditos Activos',
        'applications_with_pending_documents': 'Reporte de Solicitudes con Documentos Pendientes',
        'verifications_by_status': 'Reporte de Verificaciones por Estado',
        'tenants_by_status': 'Reporte de Tenants por Estado',
        'users_by_tenant': 'Reporte de Usuarios por Tenant',
        'subscriptions_by_status': 'Reporte de Suscripciones por Estado',
    }
    
    def __init__(self):
        """Inicializa el servicio de exportación."""
        pass
    
    def export(
        self,
        data: List[Dict[str, Any]],
        columns: List[str],
        format: str,
        report_metadata: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Exporta datos a formato especificado.
        
        Args:
            data: Lista de diccionarios con datos
            columns: Lista de columnas a incluir
            format: Formato de exportación (csv, xlsx, pdf)
            report_metadata: Metadatos del reporte (título, usuario, filtros, etc.)
        
        Returns:
            Contenido del archivo en bytes
        
        Raises:
            ValueError: Si el formato no es soportado
        """
        if format not in self.SUPPORTED_FORMATS:
            raise ValueError(
                f"Formato no soportado: {format}. "
                f"Formatos disponibles: {self.SUPPORTED_FORMATS}"
            )
        
        if format == 'csv':
            return self._export_csv(data, columns, report_metadata)
        elif format == 'xlsx':
            if not XLSX_AVAILABLE:
                raise ImportError(
                    "openpyxl no está instalado. Instalar con: pip install openpyxl"
                )
            return self._export_xlsx(data, columns, report_metadata)
        elif format == 'pdf':
            return self._export_pdf(data, columns, report_metadata)
    
    def _export_csv(
        self,
        data: List[Dict[str, Any]],
        columns: List[str],
        report_metadata: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Exporta datos a CSV.
        
        Args:
            data: Lista de diccionarios con datos
            columns: Lista de columnas a incluir
            report_metadata: Metadatos del reporte
        
        Returns:
            Contenido CSV en bytes con UTF-8 BOM para Excel
        """
        output = io.StringIO()
        
        # Crear writer
        writer = csv.DictWriter(
            output,
            fieldnames=columns,
            extrasaction='ignore'
        )
        
        # Escribir encabezados formateados en español
        header_row = {col: self._get_column_label(col) for col in columns}
        writer.writerow(header_row)
        
        # Escribir filas
        for row in data:
            # Serializar valores
            serialized_row = {
                col: self._serialize_value_for_csv(row.get(col))
                for col in columns
            }
            writer.writerow(serialized_row)
        
        # Obtener contenido
        csv_content = output.getvalue()
        output.close()
        
        # Convertir a bytes con UTF-8 BOM para Excel
        return '\ufeff'.encode('utf-8') + csv_content.encode('utf-8')
    
    def _export_xlsx(
        self,
        data: List[Dict[str, Any]],
        columns: List[str],
        report_metadata: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Exporta datos a XLSX (Excel) con formato profesional.
        
        Args:
            data: Lista de diccionarios con datos
            columns: Lista de columnas a incluir
            report_metadata: Metadatos del reporte (título, usuario, filtros, etc.)
        
        Returns:
            Contenido XLSX en bytes
        """
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        current_row = 1
        
        # ===== SECCIÓN DE TÍTULO Y METADATOS =====
        if report_metadata:
            # Título del reporte
            report_type = report_metadata.get('report_type', '')
            title = self.REPORT_TITLES.get(report_type, 'Reporte de Datos')
            
            title_cell = ws.cell(row=current_row, column=1)
            title_cell.value = title
            title_cell.font = Font(bold=True, size=16, color="1F4E78")
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1
            
            # Línea separadora
            current_row += 1
            
            # Metadatos del reporte
            metadata_font = Font(size=10, color="666666")
            metadata_label_font = Font(bold=True, size=10, color="333333")
            
            # Fecha de generación
            date_label = ws.cell(row=current_row, column=1)
            date_label.value = "Fecha de Generación:"
            date_label.font = metadata_label_font
            
            date_value = ws.cell(row=current_row, column=2)
            date_value.value = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            date_value.font = metadata_font
            current_row += 1
            
            # Usuario
            if report_metadata.get('user_name'):
                user_label = ws.cell(row=current_row, column=1)
                user_label.value = "Generado por:"
                user_label.font = metadata_label_font
                
                user_value = ws.cell(row=current_row, column=2)
                user_value.value = report_metadata['user_name']
                user_value.font = metadata_font
                current_row += 1
            
            # Institución (para reportes TENANT)
            if report_metadata.get('tenant_name'):
                tenant_label = ws.cell(row=current_row, column=1)
                tenant_label.value = "Institución:"
                tenant_label.font = metadata_label_font
                
                tenant_value = ws.cell(row=current_row, column=2)
                tenant_value.value = report_metadata['tenant_name']
                tenant_value.font = metadata_font
                current_row += 1
            
            # Total de registros
            total_label = ws.cell(row=current_row, column=1)
            total_label.value = "Total de Registros:"
            total_label.font = metadata_label_font
            
            total_value = ws.cell(row=current_row, column=2)
            total_value.value = len(data)
            total_value.font = metadata_font
            current_row += 1
            
            # Filtros aplicados (si existen)
            if report_metadata.get('filters'):
                filters_label = ws.cell(row=current_row, column=1)
                filters_label.value = "Filtros Aplicados:"
                filters_label.font = metadata_label_font
                current_row += 1
                
                filters = report_metadata['filters']
                # Manejar tanto formato de lista como diccionario
                if isinstance(filters, list):
                    for filter_item in filters:
                        field = filter_item.get('field', '')
                        operator = filter_item.get('operator', '')
                        value = filter_item.get('value', '')
                        filter_cell = ws.cell(row=current_row, column=2)
                        filter_cell.value = f"• {self._format_filter_label(field)} {operator}: {value}"
                        filter_cell.font = Font(size=9, color="666666", italic=True)
                        current_row += 1
                elif isinstance(filters, dict):
                    for filter_key, filter_value in filters.items():
                        filter_cell = ws.cell(row=current_row, column=2)
                        filter_cell.value = f"• {self._format_filter_label(filter_key)}: {filter_value}"
                        filter_cell.font = Font(size=9, color="666666", italic=True)
                        current_row += 1
            
            # Línea separadora antes de los datos
            current_row += 1
        
        # ===== SECCIÓN DE DATOS =====
        header_row = current_row
        
        # Estilos para encabezados
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Estilos para bordes
        border_side = Side(style='thin', color="D0D0D0")
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Escribir encabezados en español
        for col_idx, column in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = self._get_column_label(column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Escribir datos
        data_start_row = header_row + 1
        for row_idx, row_data in enumerate(data, start=data_start_row):
            for col_idx, column in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(column)
                
                # Serializar y escribir valor
                cell.value = self._serialize_value_for_xlsx(value)
                cell.border = border
                
                # Aplicar formato según tipo
                self._apply_cell_format(cell, value)
                
                # Alternar color de filas para mejor legibilidad
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # Ajustar ancho de columnas
        self._auto_adjust_columns(ws, columns, data, header_row)
        
        # Congelar fila de encabezados
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()
    
    def _serialize_value_for_csv(self, value: Any) -> str:
        """
        Serializa un valor para CSV.
        
        Args:
            value: Valor a serializar
        
        Returns:
            String representando el valor
        """
        if value is None:
            return ''
        
        if isinstance(value, (datetime, date)):
            if isinstance(value, datetime):
                return value.strftime('%d/%m/%Y %H:%M:%S')
            return value.strftime('%d/%m/%Y')
        
        if isinstance(value, Decimal):
            return str(value)
        
        if isinstance(value, bool):
            return 'Sí' if value else 'No'
        
        if isinstance(value, (list, dict)):
            return str(value)
        
        return str(value)
    
    def _serialize_value_for_xlsx(self, value: Any) -> Any:
        """
        Serializa un valor para XLSX.
        
        Args:
            value: Valor a serializar
        
        Returns:
            Valor en formato apropiado para Excel
        """
        if value is None:
            return ''
        
        if isinstance(value, (datetime, date)):
            return value
        
        if isinstance(value, Decimal):
            return float(value)
        
        if isinstance(value, bool):
            return 'Sí' if value else 'No'
        
        if isinstance(value, (list, dict)):
            return str(value)
        
        return value
    
    def _apply_cell_format(self, cell, value: Any):
        """
        Aplica formato a celda según tipo de valor.
        
        Args:
            cell: Celda de openpyxl
            value: Valor de la celda
        """
        if isinstance(value, datetime):
            cell.number_format = 'DD/MM/YYYY HH:MM:SS'
            cell.alignment = Alignment(horizontal="center")
        
        elif isinstance(value, date):
            cell.number_format = 'DD/MM/YYYY'
            cell.alignment = Alignment(horizontal="center")
        
        elif isinstance(value, (int, float, Decimal)):
            if isinstance(value, Decimal) or (isinstance(value, float) and value % 1 != 0):
                cell.number_format = '#,##0.00'
            else:
                cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
        
        else:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    def _get_column_label(self, column: str) -> str:
        """
        Obtiene la etiqueta en español para una columna.
        
        Args:
            column: Nombre técnico de la columna
        
        Returns:
            Etiqueta en español
        """
        return self.COLUMN_LABELS.get(column, self._format_column_header(column))
    
    def _format_filter_label(self, filter_key: str) -> str:
        """
        Formatea el nombre de un filtro para mostrarlo.
        
        Args:
            filter_key: Clave del filtro
        
        Returns:
            Nombre formateado
        """
        filter_labels = {
            'status': 'Estado',
            'risk_level': 'Nivel de Riesgo',
            'branch_id': 'Sucursal',
            'product_id': 'Producto',
            'date_range': 'Rango de Fechas',
            'created_at': 'Fecha de Creación',
            'submitted_at': 'Fecha de Envío',
        }
        return filter_labels.get(filter_key, self._format_column_header(filter_key))
    
    def _format_column_header(self, column: str) -> str:
        """
        Formatea nombre de columna para encabezado.
        
        Convierte snake_case a Title Case.
        
        Args:
            column: Nombre de columna en snake_case
        
        Returns:
            Nombre formateado
        """
        # Convertir snake_case a Title Case
        return column.replace('_', ' ').title()
    
    def _auto_adjust_columns(
        self,
        ws,
        columns: List[str],
        data: List[Dict[str, Any]],
        header_row: int = 1
    ):
        """
        Ajusta automáticamente el ancho de columnas.
        
        Args:
            ws: Worksheet de openpyxl
            columns: Lista de columnas
            data: Datos del reporte
            header_row: Fila donde están los encabezados
        """
        for col_idx, column in enumerate(columns, start=1):
            column_letter = get_column_letter(col_idx)
            
            # Calcular ancho basado en encabezado
            max_length = len(self._get_column_label(column))
            
            # Calcular ancho basado en contenido (primeras 100 filas para performance)
            sample_data = data[:100] if len(data) > 100 else data
            for row_data in sample_data:
                value = row_data.get(column)
                if value is not None:
                    # Serializar para obtener longitud real
                    serialized = self._serialize_value_for_csv(value)
                    cell_length = len(str(serialized))
                    max_length = max(max_length, cell_length)
            
            # Aplicar ancho (con límite máximo de 50 para evitar columnas muy anchas)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _export_pdf(
        self,
        data: List[Dict[str, Any]],
        columns: List[str],
        report_metadata: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Exporta datos a PDF delegando a PDFExportService.
        
        Args:
            data: Lista de diccionarios con datos
            columns: Lista de columnas a incluir
            report_metadata: Metadatos del reporte
        
        Returns:
            Contenido PDF en bytes
        """
        from .pdf_export_service import PDFExportService
        
        pdf_service = PDFExportService()
        
        # Obtener configuración del gráfico si existe en metadata
        chart_config = None
        if report_metadata and report_metadata.get('chart_config'):
            chart_config = report_metadata['chart_config']
        
        return pdf_service.export(
            data=data,
            columns=columns,
            report_metadata=report_metadata,
            chart_config=chart_config
        )
